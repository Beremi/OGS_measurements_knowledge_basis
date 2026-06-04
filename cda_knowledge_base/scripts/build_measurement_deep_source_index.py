#!/usr/bin/env python3
"""Build searchable deep summaries for the CD-A measurement source catalogue."""

from __future__ import annotations

import csv
import hashlib
import json
import math
import os
import re
import sys
import zipfile
from collections import Counter, defaultdict
from datetime import date, datetime
from pathlib import Path
from typing import Any
from xml.etree import ElementTree as ET

from openpyxl import load_workbook
from pypdf import PdfReader


ROOT = Path(__file__).resolve().parents[2]
MEASUREMENTS = ROOT / "cda_knowledge_base" / "measurements"
SOURCE_ROOTS = [
    ROOT / "cda_knowledge_base" / "attachments",
    ROOT / "cda_knowledge_base" / "file_transfers" / "collected",
]
ARCHIVE_CATALOG = MEASUREMENTS / "archive_member_catalog.csv"
SOURCE_MANIFEST = MEASUREMENTS / "source_file_manifest.csv"

MEASUREMENT_DIRS = [
    "ert",
    "nmr",
    "permeability_pulse_tests",
    "taupe_tdr",
    "suction_relative_humidity",
    "coordinates_geometry_layout",
    "bedding_geology_structure",
    "model_projection_inputs",
    "other_hm_monitoring",
]

TEXT_EXTENSIONS = {".dat", ".txt", ".ohm", ".tx0", ".csv", ".xml", ".md", ".prj"}
OFFICE_EXTENSIONS = {".xlsx", ".pptx", ".pdf"}
ARCHIVE_INTEREST_EXTENSIONS = {
    ".xlsx",
    ".xls",
    ".csv",
    ".tsv",
    ".pptx",
    ".ppt",
    ".pdf",
    ".dat",
    ".txt",
    ".ohm",
    ".tx0",
    ".vtk",
    ".vtu",
    ".xml",
    ".py",
    ".png",
    ".jpg",
    ".jpeg",
}

KEYWORDS = [
    "ERT",
    "NMR",
    "Taupe",
    "TDR",
    "suction",
    "humidity",
    "relative humidity",
    "permeability",
    "bedding",
    "fault",
    "fracture",
    "water content",
    "saturation",
    "pressure",
    "piezometer",
    "extensometer",
    "crackmeter",
    "laser",
    "levelling",
    "OGS",
    "HERMES",
    "CD-A",
    "niche",
]


def rel(path: Path) -> str:
    return path.relative_to(ROOT).as_posix()


def markdown_link(path: Path, label: str | None = None) -> str:
    label = label or path.name
    target = path.relative_to(MEASUREMENTS).as_posix()
    if " " in target or "(" in target or ")" in target:
        return f"[{label}](<{target}>)"
    return f"[{label}]({target})"


def relative_markdown_link(from_dir: Path, target: Path, label: str | None = None) -> str:
    label = label or target.name
    target_rel = os.path.relpath(target, from_dir).replace(os.sep, "/")
    if " " in target_rel or "(" in target_rel or ")" in target_rel:
        return f"[{label}](<{target_rel}>)"
    return f"[{label}]({target_rel})"


def safe_name(path: Path) -> str:
    text = path.as_posix()
    digest = hashlib.sha1(text.encode("utf-8")).hexdigest()[:8]
    name = re.sub(r"[^A-Za-z0-9._-]+", "_", path.name)
    return f"{path.stem}_{digest}{path.suffix}"


def sha1_file(path: Path, chunk_size: int = 1024 * 1024) -> str:
    h = hashlib.sha1()
    with path.open("rb") as fh:
        while True:
            chunk = fh.read(chunk_size)
            if not chunk:
                break
            h.update(chunk)
    return h.hexdigest()


def read_csv(path: Path) -> list[dict[str, str]]:
    if not path.exists():
        return []
    with path.open(newline="", encoding="utf-8") as fh:
        return list(csv.DictReader(fh))


def write_csv(path: Path, rows: list[dict[str, Any]], fieldnames: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as fh:
        writer = csv.DictWriter(fh, fieldnames=fieldnames)
        writer.writeheader()
        for row in rows:
            writer.writerow({name: row.get(name, "") for name in fieldnames})


def write_text(path: Path, text: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(text, encoding="utf-8")


def source_file_paths() -> list[tuple[str, Path]]:
    paths: list[tuple[str, Path]] = []
    for measurement in MEASUREMENT_DIRS:
        source_root = MEASUREMENTS / measurement / "source_files"
        if not source_root.exists():
            continue
        for path in source_root.rglob("*"):
            if path.is_file():
                paths.append((measurement, path))
    return sorted(paths, key=lambda item: (item[0], item[1].as_posix()))


def build_archive_provenance() -> dict[str, list[str]]:
    provenance: dict[str, list[str]] = defaultdict(list)
    for row in read_csv(ARCHIVE_CATALOG):
        archive = row.get("archive", "")
        member = row.get("archive_member", "")
        for copy_path in row.get("local_copies_or_extracted_paths", "").split(";"):
            copy_path = copy_path.strip()
            if copy_path:
                provenance[copy_path].append(f"{archive}::{member}")
    return provenance


def build_manifest_sha() -> dict[str, str]:
    out: dict[str, str] = {}
    for row in read_csv(SOURCE_MANIFEST):
        path = row.get("relative_path", "")
        sha1 = row.get("sha1", "")
        if path and sha1:
            out[path] = sha1
    return out


def build_name_candidates() -> dict[str, list[str]]:
    candidates: dict[str, list[str]] = defaultdict(list)
    for root in SOURCE_ROOTS:
        if not root.exists():
            continue
        for path in root.rglob("*"):
            if path.is_file():
                candidates[path.name].append(rel(path))
    root_zip = ROOT / "SOTA___OGS___Mont_Terri.zip"
    if root_zip.exists():
        candidates[root_zip.name].append(rel(root_zip))
    return candidates


def keyword_hits(text: str) -> str:
    lower = text.lower()
    hits = [kw for kw in KEYWORDS if kw.lower() in lower]
    return "; ".join(hits)


def normalize_text_lines(lines: list[str]) -> list[str]:
    clean: list[str] = []
    for line in lines:
        line = re.sub(r"\s+", " ", line).strip()
        if line:
            clean.append(line)
    return clean


def clip(text: Any, limit: int = 4000) -> str:
    text = "" if text is None else str(text)
    if len(text) <= limit:
        return text
    return text[:limit].rstrip() + "..."


def extract_pdf(path: Path, out_dir: Path) -> dict[str, Any]:
    text_parts: list[str] = []
    page_count = 0
    error = ""
    try:
        reader = PdfReader(str(path))
        page_count = len(reader.pages)
        for index, page in enumerate(reader.pages, start=1):
            page_text = page.extract_text() or ""
            page_text = page_text.strip()
            if page_text:
                text_parts.append(f"\n\n--- Page {index} ---\n{page_text}")
    except Exception as exc:  # noqa: BLE001 - source files can be malformed
        error = f"{type(exc).__name__}: {exc}"
    text = "".join(text_parts).strip()
    extract_path = out_dir / "extracted_text" / f"{safe_name(path)}.txt"
    if text or error:
        write_text(extract_path, text if text else f"Extraction failed: {error}\n")
    lines = normalize_text_lines(text.splitlines())
    return {
        "kind": "pdf",
        "pages": page_count,
        "text_chars": len(text),
        "extract_path": rel(extract_path) if extract_path.exists() else "",
        "keywords": keyword_hits(text),
        "first_lines": " | ".join(lines[:8]),
        "error": error,
    }


def pptx_slide_number(path: str) -> int:
    match = re.search(r"slide(\d+)\.xml$", path)
    return int(match.group(1)) if match else 0


def extract_pptx(path: Path, out_dir: Path) -> dict[str, Any]:
    slides: list[tuple[int, list[str]]] = []
    error = ""
    try:
        with zipfile.ZipFile(path) as zf:
            slide_names = sorted(
                [name for name in zf.namelist() if name.startswith("ppt/slides/slide") and name.endswith(".xml")],
                key=pptx_slide_number,
            )
            for name in slide_names:
                xml = zf.read(name)
                root = ET.fromstring(xml)
                texts = []
                for elem in root.iter():
                    if elem.tag.endswith("}t") and elem.text:
                        texts.append(elem.text)
                slides.append((pptx_slide_number(name), normalize_text_lines(texts)))
    except Exception as exc:  # noqa: BLE001
        error = f"{type(exc).__name__}: {exc}"
    text_lines: list[str] = []
    for slide_no, lines in slides:
        text_lines.append(f"--- Slide {slide_no} ---")
        text_lines.extend(lines)
        text_lines.append("")
    text = "\n".join(text_lines).strip()
    extract_path = out_dir / "extracted_text" / f"{safe_name(path)}.txt"
    if text or error:
        write_text(extract_path, text if text else f"Extraction failed: {error}\n")
    return {
        "kind": "pptx",
        "slides": len(slides),
        "text_chars": len(text),
        "extract_path": rel(extract_path) if extract_path.exists() else "",
        "keywords": keyword_hits(text),
        "first_lines": " | ".join(normalize_text_lines(text.splitlines())[:10]),
        "error": error,
    }


def cell_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    return str(value).strip()


def compact_number(value: float) -> str:
    if value == 0:
        return "0"
    if abs(value) >= 10000 or abs(value) < 0.001:
        return f"{value:.4g}"
    return f"{value:.6g}"


def summarize_workbook(path: Path, out_dir: Path) -> dict[str, Any]:
    rows: list[dict[str, Any]] = []
    text_dump: list[str] = []
    error = ""
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            if not hasattr(ws, "iter_rows"):
                rows.append(
                    {
                        "sheet": sheet_name,
                        "max_row": "chartsheet",
                        "max_column": "chartsheet",
                        "non_empty_cells": 0,
                        "first_non_empty_row": "",
                        "first_non_empty_values": "Chartsheet/no worksheet cells; chart data are in the workbook data sheets.",
                        "date_range": "",
                        "numeric_summary": "",
                        "keywords": "",
                    }
                )
                text_dump.append(f"\n\n--- Sheet: {sheet_name} ---\nChartsheet/no worksheet cells.")
                continue
            non_empty = 0
            first_non_empty_row = None
            first_non_empty_values: list[str] = []
            numeric_by_col: dict[int, list[float]] = defaultdict(list)
            date_values: list[date | datetime] = []
            sampled_rows: list[list[str]] = []
            for row_index, row in enumerate(ws.iter_rows(values_only=True), start=1):
                values = [cell_text(value) for value in row]
                if any(values):
                    if first_non_empty_row is None:
                        first_non_empty_row = row_index
                        first_non_empty_values = values[:20]
                    if len(sampled_rows) < 30:
                        sampled_rows.append(values[:20])
                    for col_index, value in enumerate(row, start=1):
                        if value is None or value == "":
                            continue
                        non_empty += 1
                        if isinstance(value, (datetime, date)):
                            date_values.append(value)
                        elif isinstance(value, (int, float)) and math.isfinite(float(value)):
                            numeric_by_col[col_index].append(float(value))
            numeric_parts: list[str] = []
            for col_index, values in sorted(numeric_by_col.items()):
                if len(values) >= 3:
                    numeric_parts.append(
                        f"C{col_index}: n={len(values)}, min={compact_number(min(values))}, max={compact_number(max(values))}"
                    )
                if len(numeric_parts) >= 8:
                    break
            date_range = ""
            if date_values:
                dates = sorted(date_values)
                date_range = f"{cell_text(dates[0])} to {cell_text(dates[-1])} (n={len(dates)})"
            header = " | ".join(value for value in first_non_empty_values if value)[:800]
            sample_text = "\n".join("\t".join(value for value in row if value) for row in sampled_rows)
            text_dump.append(f"\n\n--- Sheet: {sheet_name} ---\n{sample_text}".strip())
            rows.append(
                {
                    "sheet": sheet_name,
                    "max_row": ws.max_row,
                    "max_column": ws.max_column,
                    "non_empty_cells": non_empty,
                    "first_non_empty_row": first_non_empty_row or "",
                    "first_non_empty_values": header,
                    "date_range": date_range,
                    "numeric_summary": "; ".join(numeric_parts),
                    "keywords": keyword_hits(sample_text),
                }
            )
        wb.close()
    except Exception as exc:  # noqa: BLE001
        error = f"{type(exc).__name__}: {exc}"
    extract_path = out_dir / "extracted_text" / f"{safe_name(path)}.workbook-outline.txt"
    if text_dump or error:
        write_text(extract_path, "\n".join(text_dump).strip() if text_dump else f"Extraction failed: {error}\n")
    return {
        "kind": "xlsx",
        "sheets": len(rows),
        "workbook_rows": rows,
        "text_chars": sum(len(item) for item in text_dump),
        "extract_path": rel(extract_path) if extract_path.exists() else "",
        "keywords": "; ".join(sorted({kw for row in rows for kw in row.get("keywords", "").split("; ") if kw})),
        "first_lines": " | ".join(row.get("first_non_empty_values", "") for row in rows[:3])[:1000],
        "error": error,
    }


def summarize_text_file(path: Path, out_dir: Path) -> dict[str, Any]:
    size = path.stat().st_size
    error = ""
    lines: list[str] = []
    line_count: int | str = ""
    try:
        if size > 80_000_000:
            with path.open("r", encoding="utf-8", errors="replace") as fh:
                for _ in range(80):
                    line = fh.readline()
                    if not line:
                        break
                    lines.append(line.rstrip("\n"))
            line_count = "not counted (large file)"
        else:
            with path.open("r", encoding="utf-8", errors="replace") as fh:
                for index, line in enumerate(fh, start=1):
                    if len(lines) < 80:
                        lines.append(line.rstrip("\n"))
                line_count = index if "index" in locals() else 0
    except Exception as exc:  # noqa: BLE001
        error = f"{type(exc).__name__}: {exc}"
    clean = normalize_text_lines(lines)
    extract_path = out_dir / "extracted_text" / f"{safe_name(path)}.head.txt"
    if clean or error:
        write_text(extract_path, "\n".join(clean) if clean else f"Extraction failed: {error}\n")
    return {
        "kind": "text/data",
        "lines": line_count,
        "extract_path": rel(extract_path) if extract_path.exists() else "",
        "keywords": keyword_hits("\n".join(clean)),
        "first_lines": " | ".join(clean[:10]),
        "error": error,
    }


def summarize_zip(path: Path, out_dir: Path, loose_names: set[str], global_loose_names: set[str]) -> dict[str, Any]:
    error = ""
    rows: list[dict[str, Any]] = []
    ext_counter: Counter[str] = Counter()
    interesting: list[str] = []
    missing_loose: list[str] = []
    total_uncompressed = 0
    try:
        with zipfile.ZipFile(path) as zf:
            for info in zf.infolist():
                if info.is_dir():
                    continue
                ext = Path(info.filename).suffix.lower()
                ext_counter[ext or "[no extension]"] += 1
                total_uncompressed += info.file_size
                if ext in ARCHIVE_INTEREST_EXTENSIONS:
                    interesting.append(info.filename)
                    if Path(info.filename).name not in global_loose_names and ext in OFFICE_EXTENSIONS | {".dat", ".png", ".jpg", ".jpeg"}:
                        missing_loose.append(info.filename)
                    rows.append(
                        {
                            "archive": rel(path),
                            "member": info.filename,
                            "extension": ext,
                            "size_bytes": info.file_size,
                            "compressed_bytes": info.compress_size,
                            "loose_file_with_same_name_present": Path(info.filename).name in loose_names,
                            "loose_file_with_same_name_present_anywhere": Path(info.filename).name in global_loose_names,
                        }
                    )
    except Exception as exc:  # noqa: BLE001
        error = f"{type(exc).__name__}: {exc}"
    csv_path = out_dir / "zip_member_deep_summary.csv"
    if rows:
        write_csv(
            csv_path,
            rows,
            [
                "archive",
                "member",
                "extension",
                "size_bytes",
                "compressed_bytes",
                "loose_file_with_same_name_present",
                "loose_file_with_same_name_present_anywhere",
            ],
        )
    return {
        "kind": "zip",
        "members": sum(ext_counter.values()),
        "total_uncompressed_bytes": total_uncompressed,
        "extension_counts": "; ".join(f"{ext}={count}" for ext, count in sorted(ext_counter.items())),
        "interesting_members": clip(" | ".join(interesting[:30])),
        "potential_missing_loose_members": clip(" | ".join(missing_loose[:30])),
        "zip_member_csv": rel(csv_path) if csv_path.exists() else "",
        "keywords": keyword_hits(" ".join(interesting)),
        "first_lines": " | ".join(interesting[:10]),
        "error": error,
    }


def provenance_for(
    path: Path,
    archive_provenance: dict[str, list[str]],
    name_candidates: dict[str, list[str]],
) -> str:
    path_rel = rel(path)
    items: list[str] = []
    for archive_item in archive_provenance.get(path_rel, []):
        items.append(f"archive member {archive_item}")
    for candidate in name_candidates.get(path.name, []):
        if candidate != path_rel and "/measurements/" not in candidate:
            items.append(candidate)
    return "; ".join(dict.fromkeys(items))


def analyze_file(
    measurement: str,
    path: Path,
    out_dir: Path,
    manifest_sha: dict[str, str],
    archive_provenance: dict[str, list[str]],
    name_candidates: dict[str, list[str]],
    loose_names: set[str],
    global_loose_names: set[str],
) -> tuple[dict[str, Any], list[dict[str, Any]]]:
    ext = path.suffix.lower()
    details: dict[str, Any] = {}
    workbook_rows: list[dict[str, Any]] = []
    if ext == ".pdf":
        details = extract_pdf(path, out_dir)
    elif ext == ".pptx":
        details = extract_pptx(path, out_dir)
    elif ext == ".xlsx":
        details = summarize_workbook(path, out_dir)
        workbook_rows = details.pop("workbook_rows", [])
    elif ext == ".zip":
        details = summarize_zip(path, out_dir, loose_names, global_loose_names)
    elif ext in TEXT_EXTENSIONS:
        details = summarize_text_file(path, out_dir)
    else:
        details = {"kind": ext.lstrip(".") or "binary", "keywords": "", "first_lines": "", "error": ""}
    path_rel = rel(path)
    sha1 = manifest_sha.get(path_rel, "")
    if not sha1 and path.stat().st_size <= 500_000_000:
        sha1 = sha1_file(path)
    row = {
        "measurement": measurement,
        "source_path": path_rel,
        "filename": path.name,
        "extension": ext,
        "size_bytes": path.stat().st_size,
        "sha1": sha1,
        "provenance": provenance_for(path, archive_provenance, name_candidates),
        **details,
    }
    for key in {
        "first_lines",
        "interesting_members",
        "potential_missing_loose_members",
        "extension_counts",
        "keywords",
        "provenance",
    }:
        if isinstance(row.get(key), str):
            row[key] = clip(row[key])
    for sheet_row in workbook_rows:
        sheet_row["measurement"] = measurement
        sheet_row["workbook"] = path_rel
    return row, workbook_rows


def build_measurement_markdown(measurement: str, rows: list[dict[str, Any]], workbook_rows: list[dict[str, Any]]) -> str:
    title = measurement.replace("_", " ").title()
    md_dir = MEASUREMENTS / measurement / "derived_files" / "deep_source_pass"
    lines = [
        f"# Deep Source Pass: {title}",
        "",
        "This generated file records the detailed pass through copied source files in this measurement folder.",
        "It is meant for fast orientation; the raw files in `source_files/` remain the authoritative data.",
        "",
        "## Source Files",
        "",
    ]
    for row in rows:
        source = ROOT / row["source_path"]
        source_link = relative_markdown_link(md_dir, source)
        lines.extend(
            [
                f"### {row['filename']}",
                "",
                f"- Local copy: {source_link}",
                f"- Type: `{row.get('kind', row.get('extension', ''))}`; size: {row.get('size_bytes', '')} bytes",
                f"- SHA1: `{row.get('sha1', '')}`" if row.get("sha1") else "- SHA1: not calculated for this pass",
            ]
        )
        provenance = row.get("provenance", "")
        if provenance:
            lines.append(f"- Original/raw provenance candidates: {provenance}")
        extract = row.get("extract_path", "")
        if extract:
            lines.append(f"- Searchable extract/outline: {relative_markdown_link(md_dir, ROOT / extract)}")
        if row.get("pages"):
            lines.append(f"- PDF pages: {row['pages']}; extracted text characters: {row.get('text_chars', 0)}")
        if row.get("slides"):
            lines.append(f"- PPTX slides: {row['slides']}; extracted text characters: {row.get('text_chars', 0)}")
        if row.get("sheets"):
            lines.append(f"- Workbook sheets: {row['sheets']}; outline characters: {row.get('text_chars', 0)}")
        if row.get("members"):
            lines.append(
                f"- ZIP members: {row['members']}; uncompressed bytes: {row.get('total_uncompressed_bytes', '')}"
            )
            if row.get("extension_counts"):
                lines.append(f"- ZIP extension counts: {row['extension_counts']}")
            if row.get("zip_member_csv"):
                lines.append(f"- ZIP member CSV: {relative_markdown_link(md_dir, ROOT / row['zip_member_csv'])}")
            if row.get("potential_missing_loose_members"):
                lines.append(
                    "- Office/data/image members without a same-name loose source copy anywhere in the "
                    f"measurement catalogue: {row['potential_missing_loose_members']}"
                )
        if row.get("lines"):
            lines.append(f"- Text/data line count: {row['lines']}")
        if row.get("keywords"):
            lines.append(f"- Keyword hits: {row['keywords']}")
        if row.get("first_lines"):
            snippet = row["first_lines"]
            if len(snippet) > 900:
                snippet = snippet[:900] + "..."
            lines.append(f"- First extracted lines / sheet headers: {snippet}")
        if row.get("error"):
            lines.append(f"- Extraction warning: {row['error']}")
        lines.append("")
    if workbook_rows:
        lines.extend(["## Workbook Sheet Details", ""])
        for row in workbook_rows:
            lines.extend(
                [
                    f"### {Path(row['workbook']).name} / {row['sheet']}",
                    "",
                    f"- Used dimensions: {row.get('max_row', '')} rows x {row.get('max_column', '')} columns",
                    f"- Non-empty cells: {row.get('non_empty_cells', '')}",
                    f"- First non-empty row: {row.get('first_non_empty_row', '')}",
                ]
            )
            if row.get("first_non_empty_values"):
                lines.append(f"- First/header values: {row['first_non_empty_values']}")
            if row.get("date_range"):
                lines.append(f"- Date range detected: {row['date_range']}")
            if row.get("numeric_summary"):
                lines.append(f"- Numeric ranges detected: {row['numeric_summary']}")
            lines.append("")
    return "\n".join(lines).rstrip() + "\n"


def build_central_markdown(rows: list[dict[str, Any]], workbook_rows: list[dict[str, Any]]) -> str:
    by_measurement: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in rows:
        by_measurement[row["measurement"]].append(row)
    kind_counts = Counter(row.get("kind", "") for row in rows)
    lines = [
        "# Measurement Deep Source Index",
        "",
        "Generated deep pass through the copied measurement source files. It indexes office documents, PDFs, archives, raw text/data files, and workbook sheets.",
        "",
        "## Coverage",
        "",
        f"- Measurement folders scanned: {len(by_measurement)}.",
        f"- Source files scanned: {len(rows)}.",
        f"- Workbook sheets summarized: {len(workbook_rows)}.",
        "- File kinds: "
        + ", ".join(f"`{kind}`={count}" for kind, count in sorted(kind_counts.items()) if kind)
        + ".",
        "",
        "Per-measurement generated detail files:",
        "",
    ]
    for measurement in MEASUREMENT_DIRS:
        if measurement not in by_measurement:
            continue
        md_path = MEASUREMENTS / measurement / "derived_files" / "deep_source_pass" / "source_file_deep_summary.md"
        csv_path = MEASUREMENTS / measurement / "derived_files" / "deep_source_pass" / "source_file_deep_summary.csv"
        lines.append(
            f"- [{measurement}](./{measurement}/derived_files/deep_source_pass/source_file_deep_summary.md): "
            f"{len(by_measurement[measurement])} files; CSV [source_file_deep_summary.csv](./{measurement}/derived_files/deep_source_pass/source_file_deep_summary.csv)."
        )
    lines.extend(["", "## Files With Extracted Searchable Text Or Outlines", ""])
    for row in rows:
        extract = row.get("extract_path", "")
        if not extract:
            continue
        source = row["source_path"]
        extract_rel = Path(extract).relative_to("cda_knowledge_base/measurements").as_posix()
        lines.append(
            f"- `{row['measurement']}`: [{Path(source).name}](./{Path(source).relative_to('cda_knowledge_base/measurements').as_posix()}) "
            f"-> [{Path(extract).name}](./{extract_rel})"
        )
    lines.extend(["", "## Archive Checks", ""])
    zip_rows = [row for row in rows if row.get("kind") == "zip"]
    for row in zip_rows:
        lines.append(f"### {row['filename']}")
        lines.append("")
        lines.append(f"- Measurement folder: `{row['measurement']}`.")
        lines.append(f"- Members: {row.get('members', '')}; extension counts: {row.get('extension_counts', '')}.")
        if row.get("interesting_members"):
            lines.append(f"- Interesting members: {row['interesting_members']}")
        missing = row.get("potential_missing_loose_members", "")
        if missing:
            lines.append(
                "- Office/data/image members without a same-name loose source copy anywhere in the "
                f"measurement catalogue: {missing}"
            )
        else:
            lines.append(
                "- No office/data/image member from this archive was identified as lacking a same-name "
                "loose source copy in the measurement catalogue."
            )
        lines.append("")
    return "\n".join(lines).rstrip() + "\n"


def main() -> int:
    archive_provenance = build_archive_provenance()
    manifest_sha = build_manifest_sha()
    name_candidates = build_name_candidates()
    all_source_paths = source_file_paths()
    loose_names_by_measurement: dict[str, set[str]] = defaultdict(set)
    global_loose_names: set[str] = set()
    for measurement, path in all_source_paths:
        loose_names_by_measurement[measurement].add(path.name)
        global_loose_names.add(path.name)

    all_rows: list[dict[str, Any]] = []
    all_workbook_rows: list[dict[str, Any]] = []
    rows_by_measurement: dict[str, list[dict[str, Any]]] = defaultdict(list)
    workbook_by_measurement: dict[str, list[dict[str, Any]]] = defaultdict(list)

    for measurement, path in all_source_paths:
        out_dir = MEASUREMENTS / measurement / "derived_files" / "deep_source_pass"
        row, workbook_rows = analyze_file(
            measurement,
            path,
            out_dir,
            manifest_sha,
            archive_provenance,
            name_candidates,
            loose_names_by_measurement[measurement],
            global_loose_names,
        )
        all_rows.append(row)
        all_workbook_rows.extend(workbook_rows)
        rows_by_measurement[measurement].append(row)
        workbook_by_measurement[measurement].extend(workbook_rows)

    fields = [
        "measurement",
        "source_path",
        "filename",
        "extension",
        "kind",
        "size_bytes",
        "sha1",
        "provenance",
        "pages",
        "slides",
        "sheets",
        "members",
        "lines",
        "total_uncompressed_bytes",
        "extension_counts",
        "keywords",
        "extract_path",
        "zip_member_csv",
        "potential_missing_loose_members",
        "first_lines",
        "error",
    ]
    workbook_fields = [
        "measurement",
        "workbook",
        "sheet",
        "max_row",
        "max_column",
        "non_empty_cells",
        "first_non_empty_row",
        "first_non_empty_values",
        "date_range",
        "numeric_summary",
        "keywords",
    ]

    write_csv(MEASUREMENTS / "deep_source_index.csv", all_rows, fields)
    write_csv(MEASUREMENTS / "workbook_sheet_deep_index.csv", all_workbook_rows, workbook_fields)
    write_csv(
        SOURCE_MANIFEST,
        [
            {
                "measurement_folder": row["measurement"],
                "relative_path": row["source_path"],
                "filename": row["filename"],
                "extension": row["extension"],
                "size_bytes": row["size_bytes"],
                "sha1": row["sha1"],
            }
            for row in all_rows
        ],
        ["measurement_folder", "relative_path", "filename", "extension", "size_bytes", "sha1"],
    )
    write_text(MEASUREMENTS / "deep_source_index.md", build_central_markdown(all_rows, all_workbook_rows))
    summary = {
        "measurement_folders_scanned": sorted(rows_by_measurement),
        "source_files_scanned": len(all_rows),
        "workbook_sheets_summarized": len(all_workbook_rows),
        "kind_counts": dict(Counter(row.get("kind", "") for row in all_rows)),
        "generated_files": [
            rel(MEASUREMENTS / "deep_source_index.md"),
            rel(MEASUREMENTS / "deep_source_index.csv"),
            rel(MEASUREMENTS / "workbook_sheet_deep_index.csv"),
            rel(SOURCE_MANIFEST),
        ],
    }
    write_text(MEASUREMENTS / "deep_source_index_summary.json", json.dumps(summary, indent=2, sort_keys=True) + "\n")

    for measurement, rows in rows_by_measurement.items():
        out_dir = MEASUREMENTS / measurement / "derived_files" / "deep_source_pass"
        write_csv(out_dir / "source_file_deep_summary.csv", rows, fields)
        if workbook_by_measurement[measurement]:
            write_csv(out_dir / "workbook_sheet_summary.csv", workbook_by_measurement[measurement], workbook_fields)
        write_text(
            out_dir / "source_file_deep_summary.md",
            build_measurement_markdown(measurement, rows, workbook_by_measurement[measurement]),
        )

    print(json.dumps(summary, indent=2, sort_keys=True))
    return 0


if __name__ == "__main__":
    sys.exit(main())
